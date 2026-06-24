"""
Xuất file Excel đối chứng: giấy tờ upload → field OCR (OpenAI) → field DS-260 → form Word.

Chạy:
  cd backend
  .venv\\Scripts\\python.exe scripts/export_ds260_mapping_excel.py

Output: backend/data/DS260_Mapping_Reference.xlsx
"""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BASE = Path(__file__).resolve().parents[1]
DATA = BASE / "data" / "doc_schemas"
OUT = BASE / "data" / "DS260_Mapping_Reference.xlsx"
AUDIT_HUNG = BASE / "_audit_hung.json"

# Nhãn trên template Word DS-260 (export_ds260.py — DS260_LABEL_PATTERNS)
WORD_FORM_LABELS: dict[str, str] = {
    "applicant_name": "Name (Last/Middle/First) — Họ và Tên",
    "applicant_name_native": "Full name in Native Language — Họ và Tên bản ngữ",
    "family_name": "Surname / Last name — Họ",
    "given_names": "Given Names — Tên",
    "date_of_birth": "Date of Birth — Ngày tháng năm sinh",
    "birth_city": "City of Birth — Thành phố nơi sinh",
    "birth_state": "State/Province of Birth — Tỉnh/Bang nơi sinh",
    "birth_country": "Country/Region of Birth — Quốc gia nơi sinh",
    "gender": "Sex — Male or Female / Nam hay Nữ",
    "current_marital_status": "Current Marital Status — Tình trạng hôn nhân hiện tại",
    "nationality": "Country/Region of Origin (Nationality) — Quốc tịch",
    "passport_number": "Passport ID — Số hộ chiếu",
    "passport_issue_date": "Passport Issuance Date — Ngày cấp hộ chiếu",
    "passport_expiration_date": "Passport Expiration Date — Ngày hết hạn hộ chiếu",
    "passport_issuing_country": "Country of Authority that Issued Passport — Quốc gia cấp",
    "current_address": "Current Address — Địa chỉ hiện tại",
    "current_city": "City — Thành phố",
    "current_state": "State/Province — Tỉnh/Bang",
    "current_country": "Country/Region — Quốc gia",
    "postal_code": "Postal Zone/Zip Code — Mã bưu điện",
    "address_from_date": "From Date — Ở từ ngày",
    "primary_phone": "Primary Phone — Điện thoại chính",
    "email": "Email Address — Địa chỉ email",
    "social_media_platform": "Social Media Provider/Platform — Mạng xã hội",
    "social_media_identifier": "Social Media Identifier — Link/tên MXH",
    "father_surname": "Father's Surnames — Họ cha",
    "father_given_names": "Father's Given Names — Tên cha",
    "father_date_of_birth": "Father's Date of Birth — Ngày sinh cha",
    "father_birth_city": "Father's City of Birth — TP sinh cha",
    "father_birth_state": "Father's State/Province of Birth — Tỉnh sinh cha",
    "father_birth_country": "Father's Country of Birth — Quốc gia sinh cha",
    "father_is_living": "Is your father still living?",
    "mother_surname": "Mother's Surnames — Họ mẹ",
    "mother_given_names": "Mother's Given Names — Tên mẹ",
    "mother_date_of_birth": "Mother's Date of Birth — Ngày sinh mẹ",
    "mother_birth_city": "Mother's City of Birth — TP sinh mẹ",
    "mother_birth_state": "Mother's State/Province of Birth — Tỉnh sinh mẹ",
    "mother_birth_country": "Mother's Country of Birth — Quốc gia sinh mẹ",
    "mother_is_living": "Is your mother still living?",
    "spouse_surname": "Spouse's Surnames — HỌ của CHỒNG/VỢ",
    "spouse_given_names": "Spouse's Given Names — TÊN của CHỒNG/VỢ",
    "spouse_date_of_birth": "Date of Birth (dd/mm/yyyy) — Ngày tháng năm sinh",
    "spouse_birth_city": "City of Birth — THÀNH PHỐ nơi sinh",
    "spouse_birth_state": "State/Province of Birth — TỈNH/BANG nơi sinh",
    "spouse_birth_country": "Country/Region of Birth — QUỐC GIA nơi sinh",
    "spouse_address": "Current Address — ĐỊA CHỈ hiện tại",
    "spouse_occupation": "Occupation — NGHỀ NGHIỆP",
    "spouse_occupation_other": "Specify other — NGÀNH NGHỀ ghi rõ",
    "spouse_marriage_date": "Date of Marriage — Ngày tháng năm KẾT HÔN",
    "spouse_marriage_city": "Marriage City — tại THÀNH PHỐ nào",
    "spouse_marriage_state": "Marriage State/Province — tại TỈNH/BANG nào",
    "spouse_marriage_country": "Marriage Country/Region — tại QUỐC GIA nào",
    "spouse_immigrating": "Is your spouse immigrating to the U.S with you? — NHẬP CƯ sang MỸ cùng bạn?",
    "previous_spouses_used": "Do you have any previous spouses?",
    "children_used": "Do you have any children?",
    "children_count": "Number of children — Bao nhiêu con",
}

DOC_LABELS: dict[str, str] = {
    "passport": "Passport / Hộ chiếu",
    "birth_certificate": "Birth certificate / Giấy khai sinh (chủ hồ sơ)",
    "birth_certificate_child": "Birth certificate child / Giấy khai sinh con",
    "marriage_certificate": "Marriage certificate / Giấy kết hôn",
    "divorce": "Divorce / Quyết định ly hôn",
    "judicial_certificate": "Judicial certificate / Lý lịch tư pháp",
    "death_certificate": "Death certificate / Giấy báo tử",
    "military_discharge": "Military discharge / Giấy xuất ngũ",
    "ds260_customer_form": "DS-260 khách khai (worksheet / ds260.pdf)",
    "address_document": "Address document / Giấy tờ địa chỉ",
    "spouse_applicant_profile": "Hồ sơ phối ngẫu (JSON profile)",
}


def _load_json(name: str) -> dict:
    with (DATA / name).open(encoding="utf-8") as f:
        return json.load(f)


def _infer_doc_type(detail: dict) -> str:
    """Chuẩn hóa doc_type từ audit (OCR đôi khi trả other)."""
    code = (detail.get("type") or "").strip()
    fn = (detail.get("filename") or "").upper()
    if code != "other":
        return code
    if "JUDICIAL" in fn or "LLTP" in fn:
        return "judicial_certificate"
    if "DIVORCE" in fn or "LY HON" in fn:
        return "divorce"
    if "DS260" in fn or "DS-260" in fn:
        return "ds260_customer_form"
    if "MARRIAGE" in fn or "KET HON" in fn:
        return "marriage_certificate"
    if "DEATH" in fn or "BAO TU" in fn:
        return "death_certificate"
    if "MILITARY" in fn or "XUAT NGU" in fn:
        return "military_discharge"
    if "BIRTH" in fn or "GKS" in fn or "Khai sinh" in fn.upper():
        return "birth_certificate"
    if "PASSPORT" in fn or "HO CHIEU" in fn:
        return "passport"
    return code


def _load_hung_ocr_samples() -> tuple[dict[str, dict[str, str]], list[dict]]:
    """OCR OpenAI thực tế — hồ sơ DANG VAN HUNG (_audit_hung.json)."""
    if not AUDIT_HUNG.is_file():
        return {}, []

    with AUDIT_HUNG.open(encoding="utf-8") as f:
        audit = json.load(f)

    applicants = audit.get("applicants") or []
    if not applicants:
        return {}, []

    docs_detail = applicants[0].get("documents_detail") or []
    by_doc: dict[str, dict[str, str]] = {}
    file_rows: list[dict] = []

    for detail in docs_detail:
        doc_type = _infer_doc_type(detail)
        sample = detail.get("sample") or {}
        merged: dict[str, str] = {}
        for k, v in sample.items():
            if v is None:
                continue
            s = str(v).strip()
            if s:
                merged[k] = s
        if doc_type not in by_doc:
            by_doc[doc_type] = {}
        for k, v in merged.items():
            by_doc[doc_type][k] = v
        file_rows.append(
            {
                "filename": detail.get("filename", ""),
                "doc_type": doc_type,
                "sample": merged,
            }
        )
    return by_doc, file_rows


def _lookup_openai_value(
    samples: dict[str, dict[str, str]],
    *,
    primary_doc: str,
    source_field: str,
    aliases: list[str],
    allowed_docs: list[str],
) -> tuple[str, str, str]:
    """Trả (key OpenAI dùng, giá trị, doc_type nguồn)."""
    keys = [source_field, *aliases]
    docs = [primary_doc] + [d for d in allowed_docs if d and d != primary_doc]
    for doc in docs:
        bucket = samples.get(doc, {})
        if not bucket:
            continue
        for key in keys:
            val = bucket.get(key, "")
            if val:
                return key, val, doc
    return "", "", ""


def _openai_keys_for_doc(doc_type: str, extract_keys_fn) -> list[str]:
    if not doc_type or doc_type == "spouse_applicant_profile":
        return []
    try:
        return extract_keys_fn(doc_type)
    except Exception:
        return []


def _word_label(field_key: str, field_label: str) -> str:
    return WORD_FORM_LABELS.get(field_key, field_label)


def _openai_json_example(key: str, value: str) -> str:
    if not key or not value:
        return ""
    esc = value.replace("\\", "\\\\").replace('"', '\\"')
    return f'fields.{key}.value = "{esc}"'


def _header_row(ws, headers: list[str], fill: str = "1F4E79") -> None:
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 36


def _autosize(ws, max_width: int = 48) -> None:
    for col in range(1, ws.max_column + 1):
        letter = get_column_letter(col)
        max_len = 0
        for row in range(1, min(ws.max_row + 1, 200)):
            val = ws.cell(row=row, column=col).value
            if val is not None:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), max_width)


def sheet_mapping(
    wb: Workbook,
    mapping: dict,
    allowed: dict[str, list[str]],
    hung_samples: dict[str, dict[str, str]],
    extract_keys_fn,
) -> None:
    ws = wb.active
    ws.title = "DS260 Mapping"
    headers = [
        "STT",
        "Mục DS-260 (Section)",
        "Tiêu đề mục",
        "Ghi chú nguồn",
        "Nhãn form Word DS-260",
        "Field key DS-260",
        "Nhãn hiển thị (Review)",
        "Loại giấy tờ chính (document)",
        "Tên giấy tờ (VN)",
        "Key OpenAI chính (JSON fields.*)",
        "Alias key OpenAI (JSON fields.*)",
        "Tất cả key OpenAI của doc_type",
        "Giá trị OpenAI — DANG VAN HUNG",
        "Key OpenAI có giá trị (HUNG)",
        "Giấy tờ nguồn giá trị (HUNG)",
        "Ví dụ cấu trúc OpenAI trả về",
        "Quy tắc derive",
        "Giấy tờ được phép bổ sung (enrich)",
        "Bắt buộc?",
        "Ghi chú export Word",
    ]
    _header_row(ws, headers)

    required = set(mapping.get("validation", {}).get("required_fields", []))
    row = 2
    stt = 0
    for sec in mapping.get("sections", []):
        sec_id = sec.get("id", "")
        sec_title = sec.get("title", "")
        sec_sub = sec.get("subtitle", "")
        for field in sec.get("fields", []):
            stt += 1
            key = field["key"]
            doc = field.get("document", "")
            derive = field.get("derive", "")
            alias_list = field.get("aliases") or []
            aliases = ", ".join(alias_list)
            allowed_docs = allowed.get(key, [])
            allowed_str = ", ".join(allowed_docs)
            source_field = field.get("field", "")
            ocr_keys = _openai_keys_for_doc(doc, extract_keys_fn)
            used_key, hung_val, hung_doc = _lookup_openai_value(
                hung_samples,
                primary_doc=doc,
                source_field=source_field,
                aliases=alias_list,
                allowed_docs=allowed_docs,
            )
            ws.append(
                [
                    stt,
                    sec_id,
                    sec_title,
                    sec_sub,
                    _word_label(key, field.get("label", "")),
                    key,
                    field.get("label", ""),
                    doc,
                    DOC_LABELS.get(doc, doc),
                    source_field,
                    aliases,
                    ", ".join(ocr_keys[:40]) + (" …" if len(ocr_keys) > 40 else ""),
                    hung_val,
                    used_key,
                    DOC_LABELS.get(hung_doc, hung_doc) if hung_doc else "",
                    _openai_json_example(used_key or source_field, hung_val),
                    derive or "",
                    allowed_str,
                    "Có" if key in required else "",
                    f"Key '{key}' → điền vào template Word DS-260 (export)",
                ]
            )
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).alignment = Alignment(vertical="top", wrap_text=True)
            row += 1
    _autosize(ws, max_width=55)


def sheet_documents(wb: Workbook, templates: dict) -> None:
    ws = wb.create_sheet("Giấy tờ upload")
    headers = [
        "Mã doc_type",
        "Tên hiển thị",
        "File mẫu",
        "File exception (_new)",
        "Mục form",
        "Field OCR",
        "Mô tả field",
        "Định dạng",
    ]
    _header_row(ws, headers, fill="2E7D32")

    row = 2
    for code, info in templates.get("types", {}).items():
        fields = info.get("fields", {})
        if not fields:
            ws.append([code, info.get("display_name", ""), info.get("sample_file", ""), info.get("exception_example", ""), info.get("form_section", ""), "", "", ""])
            row += 1
            continue
        first = True
        for fkey, fmeta in fields.items():
            ws.append(
                [
                    code if first else "",
                    info.get("display_name", "") if first else "",
                    info.get("sample_file", "") if first else "",
                    info.get("exception_example", "") if first else "",
                    info.get("form_section", "") if first else "",
                    fkey,
                    fmeta.get("label", ""),
                    fmeta.get("format", ""),
                ]
            )
            first = False
            row += 1
    _autosize(ws)


def sheet_openai_hung_detail(
    wb: Workbook,
    mapping: dict,
    allowed: dict[str, list[str]],
    hung_samples: dict[str, dict[str, str]],
) -> None:
    """Đối chiếu chi tiết: Form Word → OpenAI key → giá trị DANG VAN HUNG."""
    ws = wb.create_sheet("OpenAI ↔ DS-260 (HUNG)")
    headers = [
        "Mục DS-260",
        "Nhãn form Word",
        "Field key DS-260",
        "Giấy tờ chính",
        "Key OpenAI (fields.xxx)",
        "Giá trị OpenAI — DANG VAN HUNG",
        "File upload (HUNG)",
        "Ghi chú",
    ]
    _header_row(ws, headers, fill="4527A0")

    hung_files = {
        "passport": "PASSPORT - DANG VAN HUNG.pdf",
        "birth_certificate": "BIRTH CERTIFICATE / GKS - DANG VAN HUNG.pdf",
        "judicial_certificate": "JUDICIAL CERTIFICATE - DANG VAN HUNG.pdf",
        "divorce": "DIVORCE DECREE - DANG VAN HUNG.pdf",
        "ds260_customer_form": "DS260 - DANG VAN HUNG.pdf",
    }

    row = 2
    for sec in mapping.get("sections", []):
        sec_title = sec.get("title", "")
        for field in sec.get("fields", []):
            key = field["key"]
            doc = field.get("document", "")
            if doc == "spouse_applicant_profile":
                continue
            alias_list = field.get("aliases") or []
            source_field = field.get("field", "")
            used_key, hung_val, hung_doc = _lookup_openai_value(
                hung_samples,
                primary_doc=doc,
                source_field=source_field,
                aliases=alias_list,
                allowed_docs=allowed.get(key, []),
            )
            note = ""
            if key == "applicant_name_native" and not hung_val and hung_samples.get("passport", {}).get("full_name"):
                note = "Passport trả full_name_native rỗng; có thể dùng full_name bản ngữ từ GKS"
                hung_val = hung_samples.get("birth_certificate", {}).get("full_name", "")
                used_key = "full_name"
                hung_doc = "birth_certificate"
            if field.get("derive"):
                note = (note + " " if note else "") + f"derive={field['derive']}"

            ws.append(
                [
                    sec_title,
                    _word_label(key, field.get("label", "")),
                    key,
                    DOC_LABELS.get(doc, doc),
                    used_key or source_field,
                    hung_val,
                    hung_files.get(hung_doc, hung_files.get(doc, "")),
                    note.strip(),
                ]
            )
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).alignment = Alignment(vertical="top", wrap_text=True)
            row += 1
    _autosize(ws, max_width=52)


def sheet_openai_by_file(wb: Workbook, file_rows: list[dict]) -> None:
    """Mọi key OpenAI trả về theo từng file upload — DANG VAN HUNG."""
    ws = wb.create_sheet("OpenAI theo file (HUNG)")
    headers = [
        "File upload",
        "doc_type",
        "Key OpenAI (fields.xxx)",
        "Giá trị trả về",
        "Cấu trúc JSON OpenAI",
    ]
    _header_row(ws, headers, fill="00695C")

    row = 2
    for fr in file_rows:
        fn = fr.get("filename", "")
        doc_type = fr.get("doc_type", "")
        sample = fr.get("sample") or {}
        first = True
        for ocr_key in sorted(sample.keys()):
            val = sample[ocr_key]
            ws.append(
                [
                    fn if first else "",
                    doc_type if first else "",
                    ocr_key,
                    val,
                    _openai_json_example(ocr_key, val),
                ]
            )
            first = False
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).alignment = Alignment(vertical="top", wrap_text=True)
            row += 1
    _autosize(ws, max_width=60)


def sheet_openai_format(wb: Workbook) -> None:
    ws = wb.create_sheet("Cấu trúc OpenAI OCR")
    _header_row(ws, ["Mục", "Mô tả"], fill="37474F")
    rows = [
        (
            "Response JSON",
            '{ "document_type": "passport", "fields": { "full_name": { "value": "DANG VAN HUNG", "confidence": 0.95, "source_page": "page_1" } } }',
        ),
        (
            "Key trong fields",
            "Tên biến OpenAI trả về — cột Key OpenAI trong sheet DS260 Mapping",
        ),
        (
            "document_type",
            "passport | birth_certificate | judicial_certificate | divorce | ds260_customer_form | …",
        ),
        (
            "Ví dụ Name (Last/Middle/First)",
            "fields.full_name.value = ĐẶNG VĂN HÙNG (passport) hoặc fields.family_name + fields.given_names",
        ),
        (
            "Ví dụ Full name Native Language",
            "fields.full_name_native hoặc fields.name_native_language — thường từ passport/GKS; HUNG: ĐẶNG VĂN HÙNG (GKS)",
        ),
        (
            "Hồ sơ mẫu",
            "DANG VAN HUNG — nguồn backend/_audit_hung.json (OCR thực tế đã upload)",
        ),
    ]
    for r in rows:
        ws.append(list(r))
    _autosize(ws, max_width=95)
    ws.column_dimensions["B"].width = 95
    ws = wb.create_sheet("Luồng dữ liệu")
    _header_row(ws, ["Bước", "Mô tả"], fill="6A1B9A")
    rows = [
        ("0. OpenAI OCR", "OpenAI Vision trả JSON: document_type + fields.{key}.value — xem sheet Cấu trúc OpenAI OCR."),
        ("1. Upload", "User upload PDF/ảnh/Word → OCR phân loại doc_type → lưu ApplicantDocRecord (raw_data, form_data)."),
        ("2. Bảng theo file", "Mỗi file = 1 dòng trong bảng doc_type (passport, birth_certificate, ...)."),
        ("3. Resolve DS-260", "resolve_ds260_form(): mỗi field DS-260 lấy từ document + source_field trong sheet DS260 Mapping."),
        ("4. Enrich", "Field trống: bổ sung từ giấy tờ khác theo cột 'Giấy tờ được phép bổ sung' (whitelist)."),
        ("5. Worksheet", "ds260_customer_form: địa chỉ, SĐT, email, MXH — ưu tiên worksheet khách khai."),
        ("6. Xung đột", "Hai nguồn khác giá trị → Conflict; user chọn giá trị đúng."),
        ("7. Export Word", "export_ds260: map field key → ô/label trong file .docx template DS-260."),
        ("Luồng 1", "passport, birth_certificate, judicial, marriage: bản standard = nguồn chính; _new = đối chiếu."),
        ("Con cái", "birth_certificate_child: mỗi file 1 con; gộp tối đa 3 slot child_1..child_3."),
    ]
    for r in rows:
        ws.append(list(r))
    _autosize(ws, max_width=90)
    ws.column_dimensions["B"].width = 90


def sheet_sections_summary(wb: Workbook, mapping: dict) -> None:
    ws = wb.create_sheet("Tóm tắt mục")
    _header_row(ws, ["Section ID", "Tiêu đề", "Số field", "Nguồn chính"], fill="E65100")
    for sec in mapping.get("sections", []):
        docs = sorted({f.get("document", "") for f in sec.get("fields", [])})
        ws.append([sec.get("id"), sec.get("title"), len(sec.get("fields", [])), ", ".join(docs)])
    _autosize(ws)


def sheet_worksheet_fields(wb: Workbook, mapping: dict) -> None:
    """Field chỉ từ DS-260 khách khai (worksheet)."""
    ws = wb.create_sheet("Worksheet khách khai")
    headers = [
        "Field key DS-260",
        "Nhãn",
        "Field nguồn (worksheet)",
        "Alias OCR",
        "Mục DS-260",
    ]
    _header_row(ws, headers, fill="1565C0")
    for sec in mapping.get("sections", []):
        for field in sec.get("fields", []):
            if field.get("document") != "ds260_customer_form":
                continue
            ws.append(
                [
                    field["key"],
                    field.get("label", ""),
                    field.get("field", ""),
                    ", ".join(field.get("aliases") or []),
                    sec.get("title", ""),
                ]
            )
    _autosize(ws)


def sheet_spouse_b3(wb: Workbook, mapping: dict, allowed: dict[str, list[str]]) -> None:
    """Đối chiếu mục THÔNG TIN CỦA NGƯỜI PHỐI NGẪU (B.3) — nguồn giấy tờ."""
    ws = wb.create_sheet("Phối ngẫu (B.3)")
    headers = [
        "STT",
        "Nhãn form DS-260 (EN/VN)",
        "Field key",
        "Giấy tờ chính (upload)",
        "Key OpenAI chính",
        "Alias OpenAI",
        "Bổ sung nếu thiếu (enrich)",
        "Ghi chú",
    ]
    _header_row(ws, headers, fill="AD1457")

    spouse_notes = {
        "spouse_surname": "Lấy họ vợ/chồng (không trùng chủ hồ sơ) từ giấy kết hôn",
        "spouse_given_names": "Lấy tên vợ/chồng từ giấy kết hôn",
        "spouse_date_of_birth": "wife_date_of_birth / husband_date_of_birth trên giấy kết hôn",
        "spouse_birth_city": "Ưu tiên giấy kết hôn; nếu trống → GKS phối ngẫu (birth_certificate)",
        "spouse_birth_state": "Ưu tiên giấy kết hôn; nếu trống → GKS phối ngẫu",
        "spouse_birth_country": "Ưu tiên giấy kết hôn; nếu trống → GKS phối ngẫu",
        "spouse_address": "Địa chỉ hiện tại trên giấy kết hôn (wife_address / husband_address)",
        "spouse_occupation": "Ưu tiên giấy kết hôn; nếu trống → hồ sơ DS-260 của phối ngẫu",
        "spouse_occupation_other": "Ghi rõ ngành nghề — giấy kết hôn hoặc worksheet phối ngẫu",
        "spouse_marriage_date": "marriage_date trên giấy kết hôn",
        "spouse_marriage_city": "marriage_place / marriage_city — tách thành phố",
        "spouse_marriage_state": "marriage_state hoặc tách từ marriage_place",
        "spouse_marriage_country": "marriage_country hoặc tách từ marriage_place",
        "spouse_immigrating": "KHÔNG có trên giấy kết hôn — khách khai trên worksheet DS-260 (Yes/No)",
    }

    sec = next((s for s in mapping.get("sections", []) if s.get("id") == "section_spouse"), None)
    if not sec:
        return

    row = 2
    stt = 0
    for field in sec.get("fields", []):
        key = field["key"]
        if key in ("marriage_husband_name", "marriage_wife_name", "marriage_document_number", "spouse_full_name"):
            continue
        stt += 1
        doc = field.get("document", "")
        enrich = [d for d in allowed.get(key, []) if d and d != doc]
        enrich_str = ", ".join(DOC_LABELS.get(d, d) for d in enrich)
        if key in ("spouse_occupation", "spouse_occupation_other") and "spouse_applicant_profile" not in enrich_str:
            enrich_str = (enrich_str + "; " if enrich_str else "") + "Hồ sơ phối ngẫu (spouse_applicant_profile)"
        ws.append(
            [
                stt,
                _word_label(key, field.get("label", "")),
                key,
                DOC_LABELS.get(doc, doc),
                field.get("field", ""),
                ", ".join(field.get("aliases") or []),
                enrich_str,
                spouse_notes.get(key, ""),
            ]
        )
        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).alignment = Alignment(vertical="top", wrap_text=True)
        row += 1
    _autosize(ws, max_width=58)


def sheet_by_document(wb: Workbook, mapping: dict) -> None:
    """Nhóm theo loại giấy tờ upload — dễ đối chiếu từng file."""
    ws = wb.create_sheet("Theo loại giấy tờ")
    headers = [
        "Loại giấy tờ",
        "Tên VN",
        "Field DS-260 điền từ giấy này",
        "Field OCR nguồn",
        "Mục DS-260",
    ]
    _header_row(ws, headers, fill="C62828")
    for sec in mapping.get("sections", []):
        for field in sec.get("fields", []):
            doc = field.get("document", "")
            if doc in ("spouse_applicant_profile",):
                continue
            ws.append(
                [
                    doc,
                    DOC_LABELS.get(doc, doc),
                    field["key"],
                    field.get("field", ""),
                    sec.get("title", ""),
                ]
            )
    _autosize(ws)


# --- Bảng đặt tên file upload: 01 chủ hồ sơ · 02 vợ · 03–06 con ---

FAMILY_UPLOAD_MEMBERS: list[dict[str, str]] = [
    {"code": "01", "role": "Chủ hồ sơ", "name": "DANG VAN HUNG", "kind": "adult"},
    {"code": "02", "role": "Vợ / chồng", "name": "MAI THI HUONG", "kind": "adult"},
    {"code": "03", "role": "Con 1", "name": "DANG MAI PHUONG THAO", "kind": "child"},
    {"code": "04", "role": "Con 2", "name": "DANG KHOI NGUYEN", "kind": "child"},
    {"code": "05", "role": "Con 3", "name": "TEN CON 3", "kind": "child"},
    {"code": "06", "role": "Con 4", "name": "TEN CON 4", "kind": "child"},
]

UPLOAD_SLOTS_ADULT: list[tuple[int, str, str, str]] = [
    (1, "Giấy khai sinh", "BIRTH CERTIFICATE", "birth_certificate"),
    (2, "Hộ chiếu", "PASSPORT", "passport"),
    (3, "Quyết định ly hôn", "DIVORCE DECREE", "divorce"),
    (4, "Lý lịch tư pháp", "JUDICIAL CERTIFICATE", "judicial_certificate"),
]

UPLOAD_SLOTS_CHILD: list[tuple[int, str, str, str]] = [
    (1, "Giấy khai sinh con", "BIRTH CERTIFICATE CHILD", "birth_certificate_child"),
    (2, "Hộ chiếu", "PASSPORT", "passport"),
    (3, "Quyết định ly hôn (nếu có)", "DIVORCE DECREE", "divorce"),
    (4, "Lý lịch tư pháp", "JUDICIAL CERTIFICATE", "judicial_certificate"),
]

UPLOAD_SLOTS_EXTRA: list[tuple[int, str, str, str]] = [
    (5, "Giấy kết hôn", "MARRIAGE CERTIFICATE", "marriage_certificate"),
    (6, "DS-260 khách khai", "DS260", "ds260_customer_form"),
    (7, "Giấy tờ khác", "DOCUMENT", ""),
]


def _upload_filename(code: str, seq: int, part: str, name: str) -> str:
    return f"{code}_{seq} {part} - {name}.pdf"


def sheet_upload_file_naming(wb: Workbook) -> None:
    """Bảng mapping đặt tên file — chủ hồ sơ, vợ, con 1–4 (đối chiếu upload)."""
    ws = wb.create_sheet("Đặt tên file upload", 1)
    headers = [
        "Mã người",
        "Vai trò",
        "Họ tên (ví dụ)",
        "STT file",
        "Mã file",
        "Loại giấy",
        "doc_type (hệ thống)",
        "Tên file gợi ý",
        "Ghi chú",
    ]
    _header_row(ws, headers, fill="1565C0")

    row = 2
    for member in FAMILY_UPLOAD_MEMBERS:
        code = member["code"]
        slots = UPLOAD_SLOTS_CHILD if member["kind"] == "child" else UPLOAD_SLOTS_ADULT
        first = True
        for seq, label, part, doc_type in slots:
            note = ""
            if member["kind"] == "child" and seq == 1:
                note = "Bắt buộc có từ CHILD trong tên file"
            if code == "02" and seq == 1:
                note = "GKS phối ngẫu — BIRTH CERTIFICATE (không CHILD)"
            ws.append(
                [
                    code if first else "",
                    member["role"] if first else "",
                    member["name"] if first else "",
                    seq,
                    f"{code}_{seq}",
                    label,
                    doc_type,
                    _upload_filename(code, seq, part, member["name"]),
                    note,
                ]
            )
            first = False
            for c in range(1, len(headers) + 1):
                ws.cell(row=row, column=c).alignment = Alignment(vertical="top", wrap_text=True)
            row += 1

    # Dòng phân cách — file bổ sung
    ws.append(["", "", "", "", "", "— File bổ sung (_5, _6, _7…) —", "", "", ""])
    row += 1

    for seq, label, part, doc_type in UPLOAD_SLOTS_EXTRA:
        ws.append(
            [
                "01",
                "Chủ hồ sơ (ví dụ)",
                "DANG VAN HUNG",
                seq,
                f"01_{seq}",
                label,
                doc_type or "(tùy loại)",
                _upload_filename("01", seq, part, "DANG VAN HUNG"),
                "Áp dụng tương tự 02_5, 03_6… cho từng người",
            ]
        )
        for c in range(1, len(headers) + 1):
            ws.cell(row=row, column=c).alignment = Alignment(vertical="top", wrap_text=True)
        row += 1

    _autosize(ws, max_width=52)


def sheet_family_ds260_members(wb: Workbook) -> None:
    """Đối chiếu mã người → xuất DS-260 riêng."""
    ws = wb.create_sheet("DS-260 theo người")
    headers = [
        "Mã người",
        "Vai trò",
        "Họ tên ví dụ",
        "Số file chuẩn",
        "Xuất DS-260 Word",
        "Mục DS-260 chính",
        "Ghi chú",
    ]
    _header_row(ws, headers, fill="6A1B9A")
    rows = [
        (
            "01",
            "Chủ hồ sơ",
            "DANG VAN HUNG",
            "01_1 … 01_4 (+ 01_5+)",
            "Có",
            "A.1 Personal, A.2 Passport, Father, Mother, Address, Work…",
            "Luồng 1 — passport + GKS + judicial + divorce",
        ),
        (
            "02",
            "Vợ / chồng",
            "MAI THI HUONG",
            "02_1 … 02_4",
            "Có",
            "B.3 Spouse / Phối ngẫu",
            "GKS + HC + giấy kết hôn của phối ngẫu",
        ),
        (
            "03",
            "Con 1",
            "DANG MAI PHUONG THAO",
            "03_1 … 03_4",
            "Có",
            "Personal + Passport (ẩn spouse, divorce, children)",
            "03_1 = BIRTH CERTIFICATE CHILD",
        ),
        (
            "04",
            "Con 2",
            "DANG KHOI NGUYEN",
            "04_1 … 04_4",
            "Có",
            "Personal + Passport (ẩn spouse, divorce, children)",
            "",
        ),
        (
            "05",
            "Con 3",
            "TEN CON 3",
            "05_1 … 05_4",
            "Có",
            "Giống con 1",
            "",
        ),
        (
            "06",
            "Con 4",
            "TEN CON 4",
            "06_1 … 06_4",
            "Có",
            "Giống con 1",
            "",
        ),
    ]
    for r in rows:
        ws.append(list(r))
    _autosize(ws, max_width=55)


def sheet_upload_rules(wb: Workbook) -> None:
    ws = wb.create_sheet("Quy tắc đặt tên")
    _header_row(ws, ["Quy tắc", "Chi tiết"], fill="37474F")
    rules = [
        ("Format", "{mã người}_{số file} {LOẠI GIẤY} - {HỌ TÊN}.pdf"),
        ("Mã người", "01 = chủ hồ sơ · 02 = vợ/chồng · 03–06 = con 1–4"),
        ("File chuẩn", "_1 GKS · _2 HC · _3 ly hôn · _4 lý lịch tư pháp"),
        ("File thêm", "_5 giấy kết hôn · _6 DS-260 khách · _7+ giấy khác"),
        ("Hồ sơ đơn", "Không khai báo vợ/con → tất cả file gán 01_x tự động"),
        ("Hồ sơ gia đình", "Khai báo thành viên trên Review → mã 01–06 theo vai trò"),
        ("Giấy kết hôn chung", "Có thể gán 01_5 hoặc không prefix — hệ thống gán chủ hồ sơ"),
        ("Con — GKS", "Dùng BIRTH CERTIFICATE CHILD (có từ child), không dùng GKS người lớn"),
        ("Trùng loại giấy", "File thứ 2 cùng loại → _5, _6… (theo thứ tự upload)"),
        ("Đối chiếu app", "Cột Mã file trên trang Upload hiển thị 01_1, 02_3…"),
    ]
    for r in rules:
        ws.append(list(r))
    _autosize(ws, max_width=90)
    ws.column_dimensions["B"].width = 90


OUT_NEW = BASE.parent / "DS260_Mapping_Reference_new.xlsx"


def main() -> None:
    mapping = _load_json("ds260_mapping.json")
    templates = _load_json("standard_templates.json")

    import sys

    sys.path.insert(0, str(BASE))
    from app.services.ds260_field_allowed_docs import field_allowed_docs_public
    from app.services.ds260_mapping import get_extract_keys_for_doc_type

    allowed = field_allowed_docs_public()
    hung_samples, hung_files = _load_hung_ocr_samples()

    wb = Workbook()
    sheet_mapping(wb, mapping, allowed, hung_samples, get_extract_keys_for_doc_type)
    sheet_upload_file_naming(wb)
    sheet_family_ds260_members(wb)
    sheet_upload_rules(wb)
    sheet_openai_hung_detail(wb, mapping, allowed, hung_samples)
    sheet_openai_by_file(wb, hung_files)
    sheet_openai_format(wb)
    sheet_documents(wb, templates)
    sheet_sections_summary(wb, mapping)
    sheet_spouse_b3(wb, mapping, allowed)
    sheet_worksheet_fields(wb, mapping)
    sheet_by_document(wb, mapping)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    wb.save(OUT_NEW)

    root_copy = BASE.parent / "DS260_Mapping_Reference.xlsx"
    try:
        wb.save(root_copy)
        root_msg = str(root_copy)
    except PermissionError:
        root_msg = f"{root_copy} (locked — dùng {OUT_NEW.name})"

    field_count = sum(len(s.get("fields", [])) for s in mapping.get("sections", []))
    print(f"Created: {OUT}")
    print(f"New:     {OUT_NEW}")
    print(f"Copy:    {root_msg}")
    print(f"DS-260 fields: {field_count}")
    print(f"HUNG OCR doc types: {', '.join(sorted(hung_samples.keys())) or '(none — add _audit_hung.json)'}")


if __name__ == "__main__":
    main()
